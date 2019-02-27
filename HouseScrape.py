import re
from selenium import webdriver
import bs4
import openpyxl
from openpyxl.styles import Font
import requests
from openpyxl.drawing.image import Image
from PIL import Image, ImageTk
from tkinter import Label, Button, Entry, Tk, S, N, CENTER, StringVar, PhotoImage


'''Creating GUI'''
def click(): #function to activate when button clicked
    entered_url = entry.get()

    # Accessing house listing
    try: 
        res = requests.get(entered_url)
    except:
        return Label(window, text="Hmm... Are you sure that was a URL? (Has to start with https://)", background='#FBEAEB', foreground='#2F3C7E').grid(row=5, column=0)

    Label(window, text="Information successfully scraped from ", background='#FBEAEB', foreground='#2F3C7E').grid(row=5, column=0)
    Label(window, text=entered_url, background='#FBEAEB', foreground='#2F3C7E').grid(row=6, column=0)

    # Screenshoting listing
    driver = webdriver.Chrome(executable_path='./dependencies/chromedriver/chromedriver.exe')
    driver.get(entered_url)
    driver.save_screenshot("./resources/House.png")

    # Adjusting image for compatibility with excel (replaces screnshoted image with new one, then creates a formatted version with row name)
    old_house_pic = "./resources/House.png" 
    house_pic = Image.open(old_house_pic)
    new_house_pic = house_pic.resize((320, 193), Image.ANTIALIAS)
    formatted_house_pic = new_house_pic.convert('RGB')
    formatted_house_pic.save("./resources/formatted_house" + str(sheet.max_row) + ".jpg")
    driver.quit()
    
    
    #Scraping all necessary info from pure html 
    pure_html = bs4.BeautifulSoup(res.text, "html.parser")
    result = (pure_html.get_text())

    try:
        priceREG = re.compile(r'(\$)(\d+,)?(\d{3},\d{3})(.\d+)?')
        price = priceREG.search(result)
        theprice = (price.group())
    except AttributeError:
        theprice = ('No price found')

    try:
        phoneREG = re.compile(r'((\d{3})|(\(\d{3}\)))(\s|\-|\.)(\d{3})(\s|\-|\.)(\d{4})')
        phone = phoneREG.search(result)
        thephone = (phone.group())
    except AttributeError:
        thephone = ('No phone found')

    try:
        addressREG = re.compile(r'(\d{1,4})(\s)([a-zA-Z])+(\s)([a-zA-Z]+)(.|,)(\s)([a-zA-Z]+)')
        address = addressREG.search(result)
        theaddress = (address.group())
    except AttributeError:
        theaddress = ('No address found')


    try:
        bedsREG1 = re.compile(r'(Bedrooms|beds|bd|bds|bed|bedroom|bedrooms|Bdrms|Bdrm)(:)?(\s)(\d)(\+\d)?')
        beds1 = bedsREG1.search(result)
        thebeds = (beds1.group())
    except AttributeError:
        try:
            bedsREG2 = re.compile(r'(\d)((\+)(\d))?(\s)(Bedrooms|beds|bd|bds|bed|bedroom|bedrooms|Bdrms|Bdrm)')
            beds2 = bedsREG2.search(result)
            thebeds = (beds2.group())
        except AttributeError:
            thebeds = ('No beds found')

    try:
        bathsREG1 = re.compile(r'(Baths|baths|Bathrooms|bathrooms|bth|Bth|Bath|bath)(:)?(\s)(\d)(\+\d)?')
        baths1 = bathsREG1.search(result)
        thebaths = (baths1.group())
    except AttributeError:
        try:
            bathsREG2 = re.compile(r'(\d)((\+)(\d))?(\s)(Baths|baths|Bathrooms|bathrooms|bth|Bth|Bath|bath)')
            baths2 = bathsREG2.search(result)
            thebaths = (baths2.group())
        except AttributeError:
            thebaths = ('No baths found')

    try:
        sqftREG = re.compile(r'(\d)+(\s)((S|s)q(F|f)t|(S|s)q(\.)? (F|f)t)')
        sqft = sqftREG.search(result)
        thesqft = (sqft.group())
    except AttributeError:
        thesqft = ('No sqft found')

    # Saves values as strings (only accepted as so by openpyxl)
    finaladdress = str(theaddress)
    lst_for_house_data = [str(theprice), str(thesqft), str(thebeds), str(thebaths), str(thephone)]

    # Outputs info to excel (begins with address as this skips 2 rows)
    house_image = openpyxl.drawing.image.Image("./resources/formatted_house" + str(sheet.max_row) + ".jpg")
    sheet.add_image(house_image, 'A' + str(sheet.max_row + 2))
    sheet.cell(row=(sheet.max_row + 2), column=6).font = openpyxl.styles.Font(size = 18)
    sheet.cell(row=(sheet.max_row), column=6).value = finaladdress

    # Coninues writing, but through a loop since they all skip one row
    for i in lst_for_house_data:
        sheet.cell(row=(sheet.max_row + 1), column=6).font = openpyxl.styles.Font(size = 18)
        sheet.cell(row=(sheet.max_row), column=6).value = i
    
    try:
        wb.save('./My Houses.xlsx')
        Label(window, text = "Saved onto My Houses.xlsx", background='#FBEAEB', foreground='#2F3C7E', width = 65).grid(row=7, column=0)

    except PermissionError:
        Label(window, text = "Data was not saved onto My Houses.xlsx. Did you close the file before scraping?", background='#FBEAEB', foreground='#2F3C7E').grid(row=7, column=0)


# Creating excel workbook
wb = openpyxl.Workbook()
sheet = wb.active

# GUI formatting
window = Tk()
window.configure(background='#FBEAEB')
window.title("HouseScrape")
img = Image.open("./resources/houselogo.gif")
window.iconbitmap("./resources/home-146585_1280.ico")
img = img.resize((100,100), Image.ANTIALIAS)
photoImg =  ImageTk.PhotoImage(img)
Label(window, image = photoImg, background='#FBEAEB').grid (row=0, column=0, sticky=N)
Label(window, text="HouseScrape", font=("Arial Rounded MT", 30, "bold"), background='#FBEAEB', foreground='#2F3C7E').grid (row=1, column=0, sticky=N)
Label(window, text="Please enter a URL of an online house listing", background='#FBEAEB', foreground='#2F3C7E', anchor = CENTER).grid (row=2, column=0)
entry = Entry(window, width = 48)
entry.grid(row=3, column=0)
Button(window, text="Scrape", command=click, background='#2F3C7E', foreground='#FBEAEB', width=40).grid(row=4, column=0)


window.mainloop()


